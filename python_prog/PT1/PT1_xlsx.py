import openpyxl
from openpyxl.chart import BarChart, Reference
import pandas as pd


class XLSXHelper:
    def a(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        sheet['A1'] = 'Новое значение'
        sheet['B2'] = None
        sheet['C3'] = 'Измененное значение'
        workbook.save(f'new_{file}')

    def b(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        sheet.row_dimensions[1].height = 30
        sheet.column_dimensions['A'].width = 15
        sheet.merge_cells('D1:E1')
        sheet.freeze_panes = 'C4'
        workbook.save(f'new{file}')

    def c(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        chart = BarChart()

        data = Reference(sheet, min_col=6, min_row=1, max_col=6, max_row=6)

        chart.add_data(data, titles_from_data=True)

        sheet.add_chart(chart, "E10")

        workbook.save(f'new{file}')

    def write_data(self):
        with open('data.txt', 'w') as txt_file:
            txt_file.write("Заголовок1\tЗаголовок2\tЗаголовок3\n")
            txt_file.write("Значение1\tЗначение2\tЗначение3\n")
            txt_file.write("Значение4\tЗначение5\tЗначение6\n")

    def txt_to_xlsx(self, file, output_file):
        df = pd.read_csv(file, sep='\t')
        writer = pd.ExcelWriter(output_file)
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        writer.close()

    def xlsx_to_txt(self, file, output_file):
        df = pd.read_excel(file, sheet_name='Sheet1')
        df.to_csv(output_file, sep='\t', index=False)

    def d(self):
        self.write_data()
        self.txt_to_xlsx('data.txt', 'data.xlsx')
        self.xlsx_to_txt('data.xlsx', 'data_converted.txt')

if __name__ == '__main__':
    x = XLSXHelper()

